#!/usr/bin/env python3
"""
Parse the WATV Radio NOTES sheet from the radio spreadsheet and generate:
  - Intra Support Files/radio_programs.csv (full parsed rows)
  - docs/radio-schedule.txt (mapped from the CSV: slot-1 programs, cleaned titles, newest first)

Usage:
    python update_radio_programs.py \
        --xlsx "/Volumes/Office/Public/01 TV-Radio Spreadsheets/Warning! Radio - Air Date Record.xlsx" \
        --csv "Intra Support Files/radio_programs.csv" \
        --schedule docs/radio-schedule.txt

    # Regenerate schedule only from an existing CSV:
    python update_radio_programs.py --schedule-only

    # Preview:
    python update_radio_programs.py --xlsx ... --dry-run

Output CSV columns (one row per program entry):
    week_start,week_end,day,slot,program_number,title,recorded_date,week_str,source_row

- week_start/end: ISO dates parsed from the week header (e.g. 2026-06-22)
- day: Monday..Friday
- slot: "1", "2", or "filler"
- program_number: e.g. 1147R1 , 1147SW , 1147Sun (as stored)
- title: original title from sheet (stripped)
- recorded_date: extracted trailing date from title if present (YYYY-MM-DD) or empty
- week_str: original week header
- source_row: excel row for debugging

The parser focuses on the "live" mini-table section starting ~row 9039.
It automatically stops when it hits the first week header beginning with "2020/"
(the repeated template section described in WorkToDo/Radio Program Parsing.MD).
This ensures new data appended at the end of the live section is included
without any code changes to row limits. Older template data is ignored.
"""

import argparse
import csv
import html
import io
import re
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    raise

import sys

PACIFIC = ZoneInfo("America/Los_Angeles")

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
DAY_OFFSETS = {"Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3, "Friday": 4}
DAY_ABBRS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
WEEK_HEADER_PARTS_RE = re.compile(r"^(\d{4}/\d{2}/\d{2})-(\d{4}/\d{2}/\d{2})$")


def parse_week_header_parts(week_str: str) -> tuple[str, str] | None:
    """Parse week header dates after removing all whitespace (handles '05/25- 05/29', etc.)."""
    compact = re.sub(r"\s+", "", (week_str or "").strip())
    m = WEEK_HEADER_PARTS_RE.match(compact)
    if not m:
        return None
    return m.group(1), m.group(2)


def normalize_week_str(week_str: str) -> str:
    """Normalize week headers to 'YYYY/MM/DD - YYYY/MM/DD'."""
    parts = parse_week_header_parts(week_str)
    if not parts:
        return (week_str or "").strip()
    return f"{parts[0]} - {parts[1]}"


def is_week_header(col1: str) -> bool:
    return parse_week_header_parts(col1) is not None


def parse_week_range(week_str: str):
    """Parse '2026/06/22 - 2026/06/26' -> (start_date, end_date) as datetime."""
    parts = parse_week_header_parts(week_str)
    if not parts:
        return None, None
    try:
        def to_dt(s):
            y, mo, d = map(int, s.split("/"))
            return datetime(y, mo, d)
        return to_dt(parts[0]), to_dt(parts[1])
    except Exception:
        return None, None


def week_year(week_header: str) -> int:
    try:
        return int((week_header or "").split("/")[0])
    except Exception:
        return 0


def extract_recorded_date(title: str):
    """Look for trailing date like 06/10/2026 or 6/10/26 at end of title. Return YYYY-MM-DD or ''."""
    if not title:
        return ""
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{2,4})\s*$", title)
    if not m:
        return ""
    mm, dd, yycap = m.groups()
    yyi = int(yycap)
    if len(yycap) == 4:
        year = yyi
    else:
        year = 2000 + yyi if yyi < 50 else 1900 + yyi
    try:
        dt = datetime(year, int(mm), int(dd))
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return ""


def clean_title_for_display(title: str) -> str:
    """Remove trailing recorded date and any rerun program number suffix for display use."""
    if not title:
        return ""
    t = html.unescape(title.strip())
    # Repeatedly strip trailing recorded dates and rerun program numbers (e.g. 1003R5, 1145SW).
    # Loop handles combos like "02/06/2026  1129R5" or "1095R5 1096R5 1097R2".
    prog_suffix = re.compile(r"\s+\d+(?:R[1-5]|SW|Sun)\s*$", re.IGNORECASE)
    while True:
        prev = t
        t = re.sub(r"\s+\d{1,2}/\d{1,2}/\d{2,4}\s+[ap]m\s*$", "", t, flags=re.IGNORECASE)
        t = re.sub(r"\s+\d{1,2}/\d{1,2}/\d{2,4}\s*$", "", t)
        t = prog_suffix.sub("", t)
        if t == prev:
            break
    return t.strip()


def parse_radio_sheet(xlsx_path: Path, start_row: int = 9030, end_row: int = None):
    """Yield dicts for each program entry in the live mini-table region.

    Includes weeks from 2023 onward (per the spec in WorkToDo/Radio Program Parsing.MD).
    Skips week headers with year < 2023 (older history or template blocks).
    The start_row is stable; we do not rely on a hardcoded end row.
    """
    wb = load_workbook(xlsx_path, data_only=False)
    ws = wb["WATV Radio NOTES"]

    days_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    current_week = None
    current_week_dates = (None, None)
    row = start_row
    max_row = end_row or ws.max_row

    while row <= max_row:
        col1 = ws.cell(row=row, column=1).value
        if col1 and isinstance(col1, str):
            col1 = col1.strip()

            # Detect week header (allow flexible spacing around '-'; sheet sometimes has '05/25- 05/29')
            if is_week_header(col1):
                normalized_week = normalize_week_str(col1)
                if week_year(normalized_week) < 2023:
                    row += 1
                    continue
                current_week = normalized_week
                current_week_dates = parse_week_range(normalized_week)
                row += 1
                continue

            # Detect day header
            if col1 in days_order and current_week:
                day = col1
                row += 1
                # Next 1-3 rows are usually Program 1, Program 2, Filler
                for _ in range(5):  # safety
                    if row > ws.max_row or (end_row and row > end_row):
                        break
                    slot_name = ws.cell(row=row, column=1).value
                    if slot_name and isinstance(slot_name, str):
                        slot_name = slot_name.strip()
                    prog_num = ws.cell(row=row, column=3).value
                    title = ws.cell(row=row, column=4).value

                    if slot_name in ("Program 1", "Program 2", "Filler"):
                        prog_str = str(prog_num).strip() if prog_num else ""
                        title_str = str(title).strip() if title else ""
                        if not (prog_str or title_str):
                            row += 1
                            continue  # skip completely empty slots for a cleaner CSV
                        recorded = extract_recorded_date(title_str)

                        yield {
                            "week_str": current_week,
                            "week_start": current_week_dates[0].strftime("%Y-%m-%d") if current_week_dates[0] else "",
                            "week_end": current_week_dates[1].strftime("%Y-%m-%d") if current_week_dates[1] else "",
                            "day": day,
                            "slot": "1" if slot_name == "Program 1" else ("2" if slot_name == "Program 2" else "filler"),
                            "program_number": prog_str,
                            "title": title_str,
                            "recorded_date": recorded,
                            "source_row": row,
                        }
                        row += 1
                    else:
                        # next day or week header
                        break
                continue

        row += 1


def air_date_for_entry(entry: dict) -> datetime | None:
    """Compute the calendar air date from week_start + weekday offset."""
    week_start = entry.get("week_start") or ""
    day = entry.get("day") or ""
    if not week_start or day not in DAY_OFFSETS:
        return None
    try:
        start = datetime.strptime(week_start, "%Y-%m-%d")
        return start + timedelta(days=DAY_OFFSETS[day])
    except Exception:
        return None


def format_schedule_line(air_date: datetime, title: str) -> str:
    """Format like the website/static list: 'Fri, Jul 3, 2026: Title'."""
    day_abbr = DAY_ABBRS[air_date.weekday()]
    month = MONTHS[air_date.month - 1]
    safe_title = html.escape(title, quote=False)
    return f"{day_abbr}, {month} {air_date.day}, {air_date.year}: {safe_title}"


CSV_FIELDNAMES = [
    "week_start", "week_end", "day", "slot", "program_number",
    "title", "recorded_date", "week_str", "source_row",
]


def load_radio_programs_csv(csv_path: Path) -> list[dict]:
    """Load program rows from radio_programs.csv."""
    with open(csv_path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def write_radio_programs_csv(csv_path: Path, entries: list[dict]) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDNAMES)
        writer.writeheader()
        writer.writerows(entries)


def schedule_cutoff_date(max_future_days: int) -> datetime:
    """Last air date to include: today (Pacific) + max_future_days."""
    today = datetime.now(PACIFIC).date()
    cutoff = today + timedelta(days=max_future_days)
    return datetime(cutoff.year, cutoff.month, cutoff.day)


def entries_to_schedule_lines(entries: list[dict], max_future_days: int = 10) -> list[str]:
    """Build schedule lines from CSV rows (slot 1 only, newest first)."""
    cutoff = schedule_cutoff_date(max_future_days)
    candidates: list[tuple[datetime, str]] = []

    for entry in entries:
        if entry.get("slot") != "1":
            continue
        air_date = air_date_for_entry(entry)
        if not air_date:
            continue
        if air_date > cutoff:
            continue
        title = clean_title_for_display(entry.get("title") or "")
        if not title:
            continue
        candidates.append((air_date, format_schedule_line(air_date, title)))

    candidates.sort(key=lambda x: x[0], reverse=True)
    return [line for _, line in candidates]


def write_schedule(path: Path, lines: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text("\n".join(lines) + "\n", encoding="utf-8")
    tmp.replace(path)


def generate_schedule_from_csv(csv_path: Path, max_future_days: int = 10) -> list[str]:
    """Map radio_programs.csv rows into schedule text lines."""
    rows = load_radio_programs_csv(csv_path)
    return entries_to_schedule_lines(rows, max_future_days=max_future_days)


def preview_schedule_from_entries(entries: list[dict], max_future_days: int = 10) -> list[str]:
    """Dry-run helper: round-trip entries through an in-memory CSV before mapping."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=CSV_FIELDNAMES)
    writer.writeheader()
    writer.writerows(entries)
    buf.seek(0)
    return entries_to_schedule_lines(list(csv.DictReader(buf)), max_future_days=max_future_days)


def main():
    parser = argparse.ArgumentParser(description="Parse WATV Radio NOTES sheet to clean CSV.")
    parser.add_argument(
        "--xlsx",
        default="/Volumes/Office/Public/01 TV-Radio Spreadsheets/Warning! Radio - Air Date Record.xlsx",
        help="Path to the radio spreadsheet",
    )
    parser.add_argument(
        "--csv",
        default="Intra Support Files/radio_programs.csv",
        help="Output CSV path (relative to project or absolute)",
    )
    parser.add_argument(
        "--schedule",
        default="docs/radio-schedule.txt",
        help="Output schedule text path (slot-1 programs, newest first)",
    )
    parser.add_argument("--dry-run", action="store_true", help="Parse and print sample rows only, no file written.")
    parser.add_argument(
        "--schedule-only",
        action="store_true",
        help="Skip spreadsheet parse; regenerate schedule from existing --csv only",
    )
    parser.add_argument("--start-row", type=int, default=9030, help="Approx start row for live data (stable per spec)")
    parser.add_argument("--end-row", type=int, default=None, help="Optional hard max row (for debugging); normally we stop dynamically on first 2020/ week header")
    parser.add_argument(
        "--max-future-days",
        type=int,
        default=10,
        help="Include schedule entries at most this many days beyond today (Pacific)",
    )
    args = parser.parse_args()

    project_root = Path(__file__).parent
    xlsx_path = Path(args.xlsx)
    csv_path = Path(args.csv)
    schedule_path = Path(args.schedule)
    if not csv_path.is_absolute():
        csv_path = project_root / csv_path
    if not schedule_path.is_absolute():
        schedule_path = project_root / schedule_path

    print(f"CSV path: {csv_path}")
    print(f"Schedule target: {schedule_path}")
    cutoff = schedule_cutoff_date(args.max_future_days)
    print(f"Schedule air-date cutoff: {cutoff.strftime('%Y-%m-%d')} (today PT + {args.max_future_days} days)")

    if args.schedule_only:
        if not csv_path.exists():
            print(f"ERROR: CSV not found: {csv_path}", file=sys.stderr)
            sys.exit(1)
        print(f"Mapping schedule from {csv_path}")
        schedule_lines = generate_schedule_from_csv(csv_path, max_future_days=args.max_future_days)
        print(f"Generated {len(schedule_lines)} schedule lines (slot 1).")
        if args.dry_run:
            print("\nSchedule sample (newest 8):")
            for line in schedule_lines[:8]:
                print(" ", line)
            print("\nSchedule sample (oldest 3):")
            for line in schedule_lines[-3:]:
                print(" ", line)
            return
        write_schedule(schedule_path, schedule_lines)
        print(f"Wrote {len(schedule_lines)} lines to {schedule_path}")
        print("\nNewest schedule entries:")
        for line in schedule_lines[:5]:
            print(" ", line)
        return

    print(f"Parsing {xlsx_path}")
    entries = list(parse_radio_sheet(xlsx_path, args.start_row, args.end_row))
    print(f"Parsed {len(entries)} program entries.")

    if args.dry_run:
        schedule_lines = preview_schedule_from_entries(entries, max_future_days=args.max_future_days)
        print(f"Would generate {len(schedule_lines)} schedule lines from CSV mapping (slot 1).")
        print("\nCSV sample (first 5):")
        for e in entries[:5]:
            print(e)
        print("\nCSV sample (last 3):")
        for e in entries[-3:]:
            print(e)
        print("\nSchedule sample (newest 8):")
        for line in schedule_lines[:8]:
            print(" ", line)
        print("\nSchedule sample (oldest 3):")
        for line in schedule_lines[-3:]:
            print(" ", line)
        return

    write_radio_programs_csv(csv_path, entries)
    print(f"Wrote {len(entries)} rows to {csv_path}")

    print(f"Mapping schedule from {csv_path}")
    schedule_lines = generate_schedule_from_csv(csv_path, max_future_days=args.max_future_days)
    write_schedule(schedule_path, schedule_lines)

    print(f"Wrote {len(schedule_lines)} lines to {schedule_path}")
    print("\nNewest schedule entries:")
    for line in schedule_lines[:5]:
        print(" ", line)


if __name__ == "__main__":
    main()
