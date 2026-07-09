#!/usr/bin/env python3
"""
Update docs/tv-schedule.txt from the KAZQ Excel spreadsheet.

Usage (run weekly or as needed on a machine that can see the spreadsheet):

    python update_tv_schedule.py \
        --xlsx "/Volumes/Office/Public/01 TV-Radio Spreadsheets/KAZQ.xlsx" \
        --schedule docs/tv-schedule.txt

    # Preview only:
    python update_tv_schedule.py --xlsx ... --schedule ... --dry-run

The script:
- Looks at the last usable rows of the first sheet.
- Parses the "air date" column (MMDDYY code).
- Adjusts the listed date to the actual Friday air date (the spreadsheet date is typically 2 days off / Sunday).
- Produces lines in the exact format used by tv-schedule.txt: "Jul 03, 2026: Title"
- Merges new entries (deduped) and rewrites the file with newest/future first.

Requires: pip install openpyxl   (only needed on the machine running the update)
"""

import argparse
import re
from datetime import datetime, timedelta
from pathlib import Path

import sys
try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    raise

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def parse_air_to_friday(air_val) -> datetime | None:
    """Parse MMDDYY code (str or int) and return the previous Friday."""
    if air_val is None:
        return None
    s = str(air_val).strip().zfill(6)
    if len(s) != 6 or not s.isdigit():
        return None
    try:
        mm = int(s[0:2])
        dd = int(s[2:4])
        yy = int(s[4:6])
        year = 2000 + yy if yy < 50 else 1900 + yy
        listed = datetime(year, mm, dd)
    except Exception:
        return None

    # Adjustment: listed date is usually a Sunday in the sheet; actual air = previous Friday.
    # General: back up to the nearest Friday on or before the listed date.
    dow = listed.weekday()  # 0=Mon ... 4=Fri ... 6=Sun
    days_back = (dow - 4) % 7
    friday = listed - timedelta(days=days_back)
    return friday


def format_schedule_line(friday: datetime, title: str) -> str:
    title = title.strip()
    return f"{MONTHS[friday.month-1]} {friday.day:02d}, {friday.year}: {title}"


def extract_from_excel(xlsx_path: Path, max_recent: int = 15) -> list[str]:
    """Return schedule lines (newest first) from the bottom of the first sheet."""
    wb = load_workbook(xlsx_path, data_only=False)
    ws = wb.active

    candidates: list[tuple[datetime, str]] = []

    # Scan from the bottom up, collect recent usable rows
    for row_idx in range(ws.max_row, 0, -1):
        row = [c.value for c in ws[row_idx][:5]]
        rec = row[0]
        title = row[1]
        sent = row[2]
        air = row[3]

        if rec is None or title is None:
            if candidates and len(candidates) >= 3:
                # we've passed the active block
                break
            continue

        # Column C is the sent date; a program without one hasn't been
        # supplied to the TV station yet, so keep it out of the schedule.
        if sent is None:
            continue

        try:
            rec = int(rec)
        except (ValueError, TypeError):
            continue

        if not isinstance(title, str) or not title.strip():
            continue

        friday = parse_air_to_friday(air)
        if not friday:
            continue

        # Only recent/future-ish entries (tune for weekly updates)
        if friday.year < 2025:
            if candidates:
                break
            continue

        line = format_schedule_line(friday, title)
        candidates.append((friday, line))

        if len(candidates) >= max_recent:
            break

    # Newest first
    candidates.sort(reverse=True, key=lambda x: x[0])
    return [line for _, line in candidates]


def load_existing(path: Path) -> list[str]:
    if not path.exists():
        return []
    lines = []
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line:
            lines.append(line)
    return lines


def parse_date_from_line(line: str) -> datetime | None:
    """Parse the leading 'Jun 26, 2026:' date from a schedule line."""
    m = re.match(r"^([A-Za-z]{3})\s+(\d{1,2}),\s+(\d{4}):", line)
    if not m:
        return None
    mon_str, day, year = m.groups()
    try:
        mon = MONTHS.index(mon_str) + 1
        return datetime(int(year), mon, int(day))
    except Exception:
        return None


def merge_schedules(existing: list[str], new_lines: list[str]) -> list[str]:
    """Union + sort newest first. Dedup by (date, title)."""
    seen = {}
    for line in existing + new_lines:
        dt = parse_date_from_line(line)
        if dt is None:
            continue
        key = (dt.date(), line.split(":", 1)[1].strip().lower())
        if key not in seen or dt > parse_date_from_line(seen[key]) or not seen.get(key):
            seen[key] = line

    merged = list(seen.values())
    merged.sort(key=lambda ln: parse_date_from_line(ln) or datetime.min, reverse=True)
    return merged


def main():
    ap = argparse.ArgumentParser(description="Merge KAZQ Excel TV schedule into tv-schedule.txt")
    ap.add_argument("--xlsx", required=True, help="Path to KAZQ.xlsx")
    ap.add_argument("--schedule", default="docs/tv-schedule.txt", help="Path to tv-schedule.txt")
    ap.add_argument("--dry-run", action="store_true", help="Show what would change, do not write")
    ap.add_argument("--max-recent", type=int, default=15, help="How many recent Excel rows to consider")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    schedule_path = Path(args.schedule).expanduser().resolve()

    print(f"Reading Excel: {xlsx_path}")
    new_lines = extract_from_excel(xlsx_path, max_recent=args.max_recent)
    print(f"Found {len(new_lines)} recent entries from Excel (newest first):")
    for ln in new_lines[:5]:
        print("  ", ln)
    if len(new_lines) > 5:
        print(f"  ... (+{len(new_lines)-5} more)")

    existing = load_existing(schedule_path)
    print(f"\nExisting schedule has {len(existing)} entries.")

    merged = merge_schedules(existing, new_lines)

    added = [ln for ln in merged if ln not in existing]
    removed = [ln for ln in existing if ln not in merged]  # should normally be empty

    print(f"\nAfter merge: {len(merged)} entries")
    if added:
        print(f"  NEW ({len(added)}):")
        for a in added:
            print("   +", a)
    else:
        print("  No new entries (already up to date).")

    if removed:
        print(f"  Would remove ({len(removed)}):")
        for r in removed[:3]:
            print("   -", r)

    if args.dry_run:
        print("\n--dry-run: not writing.")
        return

    if merged == existing:
        print("No changes needed.")
        return

    # Write atomically-ish
    schedule_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = schedule_path.with_suffix(schedule_path.suffix + ".tmp")
    tmp.write_text("\n".join(merged) + "\n", encoding="utf-8")
    tmp.replace(schedule_path)
    print(f"Updated {schedule_path}")


if __name__ == "__main__":
    main()
